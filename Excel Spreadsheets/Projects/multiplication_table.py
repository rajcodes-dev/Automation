"""
multiplication_table.py - program which display multiplication table
on excel from command line arguments.
"""
import sys, openpyxl
from openpyxl.styles import Font

try:
    if len(sys.argv) < 2:
        print("Usage: py multiplication_table.py <number>")
        sys.exit()

    TABLE = int(sys.argv[1])
    if TABLE < 1:
        print("Please provide number greater than 0.")
        sys.exit()
except ValueError:
    print("Error: Argument must be a integer")
    sys.exit()

BOLD_FONT = Font(name='Times New Roman', bold=True)

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet.title = 'Multiplication Table'

for num in range(2, TABLE + 2):
    sheet.cell(row=num, column=1).value = num - 1
    sheet.cell(row=num, column=1).font = BOLD_FONT
    sheet.cell(row=1, column=num).value = num - 1
    sheet.cell(row=1, column=num).font = BOLD_FONT

for row_num in range(2, TABLE + 2):
    for col_num in range(2, TABLE + 2):
        product = (row_num - 1) * (col_num - 1)
        sheet.cell(row=row_num, column=col_num).value = product

wb.save(F'Excel Spreadsheets/Projects/Table{TABLE}.xlsx')
