import openpyxl
import os
import sys

if len(sys.argv) != 4:
    print("Usage: python blankRowInserter.py <N> <M> <filename>")
    sys.exit()
try:
    start_row = int(sys.argv[1])
    num_rows = int(sys.argv[2])
    filename = sys.argv[3]
except ValueError:
    print("N and M must be integers.")
    sys.exit()

if not os.path.exists(filename):
    print(f"Error: File {filename} not found!")
    sys.exit()

print(f"Opening {filename}...")
wb = openpyxl.load_workbook(filename)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

print(f"Inserting {num_rows} blank row at row {start_row}...")

for row_num in range(1, sheet.max_row + 1):
    for col_num in range(1, sheet.max_column + 1):
        source_cell_value = sheet.cell(row=row_num, column=col_num).value

        if row_num < start_row:
            new_sheet.cell(row=row_num, column=col_num).value = source_cell_value
        else:
            new_sheet.cell(row=row_num + num_rows, column=col_num).value = source_cell_value

output_filename = "insert_" + filename
new_wb.save(output_filename)
print(f"Done! Saved to {output_filename}")
