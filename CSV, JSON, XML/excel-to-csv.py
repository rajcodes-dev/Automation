# excel-to-csv.py --> Convert excel file to csv file
import openpyxl
import csv
import os

for excel_file in os.listdir('.'):
    if not excel_file.endswith('.xlsx'):
        continue

    print(f"  Converting {excel_file} ....")

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except Exception as e:
        print(f"Could not load {excel_file}: {e}")
        continue

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        excel_title = excel_file[:-5]
        csv_filename = f"{excel_title}_{sheet_name}.csv"

        print(f"  Writing {csv_filename} ....")

        csv_file_obj = open(csv_filename, 'w', newline='', encoding='utf-8')
        csv_writer = csv.writer(csv_file_obj)

        for row_num in range(1, sheet.max_row + 1):
            row_data = []

            for col_num in range(1, sheet.max_column + 1):
                cell_value = sheet.cell(row=row_num, column=col_num).value
                row_data.append(cell_value)

            csv_writer.writerow(row_data)

        csv_file_obj.close()

print("Done.")
